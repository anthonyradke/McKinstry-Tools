def run():
    import streamlit as st
    import pandas as pd
    from datetime import timedelta
    from io import StringIO
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    import base64
    import zipfile
    import os
    from tempfile import NamedTemporaryFile, TemporaryDirectory
    import re

    # Add a back button and tool header
    col1, col2 = st.columns([1, 6])
    with col1:
        if st.button("← Back", help="Return to Tool Hub"):
            st.session_state.current_tool = None
            st.rerun()
    
    with col2:
        st.markdown("### 📈 Siemens BAS Data Cleaner")

    def round_timestamp(dt, interval):
        if pd.isna(dt):
            return None
        if interval == "15 min":
            discard = timedelta(
                minutes=dt.minute % 15,
                seconds=dt.second,
                microseconds=dt.microsecond
            )
            dt -= discard
            if discard >= timedelta(minutes=7.5):
                dt += timedelta(minutes=15)
            return dt
        else:  # "1 min"
            return dt.replace(second=0, microsecond=0)

    def simplify_name(full_name):
        if not isinstance(full_name, str):
            return full_name

        name = full_name.strip()

        # Step 1: Remove any URL in parentheses
        name = re.sub(r"\(https?://[^\)]*\)", "", name)

        # Step 2: If there's a .FLN_ section, use the part after it
        if ".FLN_" in name:
            fln_part = name.split(".FLN_")[-1]
            return "_".join(fln_part.split(".")[-4:])

        # Step 3: If .Points. exists, prefer the piece after that
        if ".Points." in name:
            return name.split(".Points.")[-1].split(".Value")[0]

        # Step 4: If parentheses contain a readable tag, use that
        if "(" in name and ")" in name:
            inside = name.split("(")[-1].split(")")[0]
            if not inside.startswith("http") and "." in inside:
                return inside.replace(".", "_")

        # Step 5: Fallback to last few parts before ".Value"
        parts = name.split(".")
        if "Value" in parts:
            parts = parts[:parts.index("Value")]
        return "_".join(parts[-4:])

    def ensure_unique_columns(df):
        seen = {}
        new_cols = []
        for col in df.columns:
            if col not in seen:
                seen[col] = 1
                new_cols.append(col)
            else:
                seen[col] += 1
                new_cols.append(f"{col}_{seen[col]}")
        df.columns = new_cols
        return df

    def process_file(file, log):
        try:
            df_raw = pd.read_csv(file, header=None, skiprows=[0, 2])
            clean_df = pd.DataFrame()
            for i in range(0, df_raw.shape[1], 4):
                try:
                    time_col = pd.to_datetime(df_raw.iloc[1:, i].astype(str), errors='coerce')
                    values = df_raw.iloc[1:, i + 1]
                    time_rounded = time_col.map(lambda dt: round_timestamp(dt, st.session_state.rounding_interval))
                    valid = time_rounded.notna()
                    title = simplify_name(str(df_raw.iloc[0, i]))

                    temp_df = pd.DataFrame({
                        "datetime": time_rounded[valid],
                        title: values[valid]
                    })
                    clean_df = temp_df if clean_df.empty else pd.merge(clean_df, temp_df, on="datetime", how="outer")
                except Exception as e:
                    log.append(f"Skipped a group: {e}")
                    continue

            if "datetime" not in clean_df.columns:
                return None

            clean_df = clean_df.groupby("datetime", as_index=False).first().sort_values("datetime").reset_index(drop=True)

            # Ensure continuous timestamps at the selected interval
            if st.session_state.rounding_interval == "15 min":
                # Only reindex when using 15-min rounding
                full_range = pd.date_range(start=clean_df['datetime'].min(), end=clean_df['datetime'].max(), freq='15min')
                clean_df = clean_df.set_index('datetime').reindex(full_range).rename_axis('datetime').reset_index()
            # else: do nothing in 1-min mode — keep natural timestamps after rounding to :00s

            return ensure_unique_columns(clean_df)
        except Exception as e:
            log.append(f"Failed to process file: {e}")
            return None

    def save_xlsx(df_dict, filename, mode):
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)  # Only remove default sheet if it exists

        if "master" in mode.lower():
            all_df = pd.concat(df_dict.values(), axis=0, ignore_index=True)
            all_df = all_df.groupby("datetime", as_index=False).first().sort_values("datetime").reset_index(drop=True)
            all_df = ensure_unique_columns(all_df)
            ws = wb.create_sheet("Master Sheet")
            for r in dataframe_to_rows(all_df, index=False, header=True):
                ws.append(r)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 31.00
            for cell in ws['A'][1:]:
                cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'

        if mode != "Combined into one file with master sheet":
            for name, df in df_dict.items():
                ws = wb.create_sheet(name[:31])
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 31.00
                for cell in ws['A'][1:]:
                    cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'

        tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        return tmp

    # --- Main Tool Interface ---
    
    st.markdown("""
    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                padding: 1.5rem; border-radius: 10px; margin-bottom: 2rem; color: white;">
        <h4 style="margin: 0; font-weight: 600;">🔧 Advanced CSV Processing Tool</h4>
        <p style="margin: 0.5rem 0 0 0; opacity: 0.9;">
            Clean and process Siemens Building Automation System data with intelligent formatting
        </p>
    </div>
    """, unsafe_allow_html=True)

    if 'processed_files' not in st.session_state:
        st.session_state.processed_files = {}
        st.session_state.log_output = []
        st.session_state.cleaned = False

    uploaded_files = st.file_uploader("Upload one or more CSV files", type=["csv"], accept_multiple_files=True)

    with st.container():
        cols = st.columns(3)
        out_format = cols[0].selectbox("Choose Output Format", ["xlsx", "csv"], key="out_format")
        mode = cols[1].selectbox("Output Mode", [
            "Separate sheets",
            "Combined into one file with separate sheets",
            "Combined into one file with master sheet",
            "Combined into one file with master & separate sheets"
        ], key="output_mode")

    rounding_interval = cols[2].selectbox("Rounding Interval", ["15 min", "1 min"], key="rounding_interval")

    if "Combined" in mode:
        custom_filename = st.text_input("Enter output file name (no extension)", value="Combined_File")
    else:
        custom_filename = "Combined_File"

    col1, col2 = st.columns([1, 1])
    with col1:
        clean_button = st.button("🧹 Clean Files", key="clean_button", use_container_width=True)

    with col2:
        reset_button = st.button("🔁 Reset All Fields", key="reset_button", use_container_width=True)

    if reset_button:
        st.rerun()

    if clean_button and uploaded_files:
        st.session_state.processed_files.clear()
        st.session_state.log_output.clear()

        progress_bar = st.progress(0, text="Processing files...")
        n = len(uploaded_files)

        for i, uploaded_file in enumerate(uploaded_files):
            df = process_file(uploaded_file, st.session_state.log_output)
            if df is not None:
                st.session_state.processed_files[uploaded_file.name] = df
                st.session_state.log_output.append(f"✅ {uploaded_file.name} cleaned successfully.")
            else:
                st.session_state.log_output.append(f"❌ Failed to process {uploaded_file.name}.")
            progress_bar.progress((i + 1) / n, text=f"Processing {i + 1} of {n} files...")

        progress_bar.empty()
        st.session_state.cleaned = True

    if st.session_state.get("cleaned"):
        if mode == "Separate sheets":
            if len(st.session_state.processed_files) == 1:
                single_name = list(st.session_state.processed_files.keys())[0]
                df = st.session_state.processed_files[single_name]
                if out_format == "csv":
                    csv = df.to_csv(index=False).encode()
                    filtered_name = single_name.replace(".csv", "").replace(".CSV", "") + "_FILTERED.csv"
                    st.download_button("Download File", csv, file_name=filtered_name, mime="text/csv")
                else:
                    filtered_name = single_name.replace(".csv", "").replace(".CSV", "") + "_FILTERED.xlsx"
                    tmp = save_xlsx({single_name: df}, filtered_name, mode)
                    with open(tmp.name, "rb") as f:
                        st.download_button("Download File", f.read(), file_name=filtered_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                with TemporaryDirectory() as tmpdir:
                    zip_path = os.path.join(tmpdir, "cleaned_files.zip")
                    with zipfile.ZipFile(zip_path, "w") as zipf:
                        for name, df in st.session_state.processed_files.items():
                            if out_format == "csv":
                                filtered_name = name.replace(".csv", "").replace(".CSV", "") + "_FILTERED.csv"
                                file_path = os.path.join(tmpdir, filtered_name)
                                df.to_csv(file_path, index=False)
                                zipf.write(file_path, arcname=filtered_name)
                            else:
                                filtered_name = name.replace(".csv", "").replace(".CSV", "") + "_FILTERED.xlsx"
                                tmp = save_xlsx({name: df}, filtered_name, mode)
                                zipf.write(tmp.name, arcname=filtered_name)
                    with open(zip_path, "rb") as f:
                        st.download_button("Download All Files as ZIP", f.read(), file_name="Cleaned_Files.zip", mime="application/zip")
        else:
            tmp = save_xlsx(st.session_state.processed_files, f"{custom_filename}.xlsx", mode)
            with open(tmp.name, "rb") as f:
                st.download_button(f"Download {custom_filename}.xlsx", f.read(), file_name=f"{custom_filename}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with st.expander("📋 View Console Log"):
        for line in st.session_state.log_output:
            st.text(line)

    # Tool-specific styling
    st.markdown("""
    <style>
    /* Base button style */
    .stButton > button {
        font-weight: 600;
        border-radius: 8px !important;
        padding: 0.6em 1.2em;
        transition: all 0.3s ease;
        border: 0px solid transparent !important;
    }

    /* Clean Files Button */
    div[data-testid="column"]:nth-child(1) .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
        color: white !important;
    }
    
    div[data-testid="column"]:nth-child(1) .stButton > button:hover {
        background: linear-gradient(135deg, #5a67d8 0%, #6b46c1 100%) !important;
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(102, 126, 234, 0.3);
    }

    /* Reset Button */
    div[data-testid="column"]:nth-child(2) .stButton > button {
        background-color: #e74c3c !important;
        color: white !important;
    }
    
    div[data-testid="column"]:nth-child(2) .stButton > button:hover {
        background-color: #c0392b !important;
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(231, 76, 60, 0.3);
    }

    /* Back button styling */
    div[data-testid="column"]:first-child .stButton > button {
        background-color: #6c757d !important;
        color: white !important;
        font-size: 0.9rem !important;
    }
    
    div[data-testid="column"]:first-child .stButton > button:hover {
        background-color: #545b62 !important;
        transform: translateY(-1px);
    }

    .stSelectbox * {
        cursor: default !important;
    }

    /* Disable text input behavior */
    div[data-baseweb="select"] input {
        pointer-events: none !important;
        caret-color: transparent !important;
        user-select: none !important;
    }

    /* Disable placeholder text selection */
    div[data-baseweb="select"] input::placeholder {
        color: #999 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # Tool-specific watermark
    st.markdown("""
    <style>
    #tool-watermark {
        position: fixed;
        bottom: 10px;
        right: 10px;
        font-size: 11px;
        color: #aaa;
        z-index: 9999;
    }
    #tool-watermark a {
        color: #aaa;
        text-decoration: none;
    }
    #tool-watermark a:hover {
        color: #888;
        text-decoration: underline;
    }
    </style>

    <div id="tool-watermark">
        <a href="https://www.linkedin.com/in/anthonyradke/" target="_blank">Made by Anthony Radke</a>
    </div>
    """, unsafe_allow_html=True)